import json
import os
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from robots.models import Robot
from robots.views.weekly_prod_report import weekly_prod_report, weekly_prod_report_to_excel


class WeeklyProdReportTestCase(TestCase):
    def setUp(self):
        # Create some Robot objects for testing
        Robot.objects.create(serial='R2-D2', model='R2', version='D2', created='2020-01-01')
        Robot.objects.create(serial='R2-D2', model='R2', version='D2', created='2023-10-07')
        Robot.objects.create(serial='R2-D3', model='R2', version='D3', created='2023-10-06')
        Robot.objects.create(serial='R3-D2', model='R3', version='D2', created='2023-10-07')
        Robot.objects.create(serial='R3-D3', model='R3', version='D3', created='2023-10-07')

    def test_request_post_weekly_prod_report(self):
        """
        Test that a POST request to the weekly_prod_report endpoint returns the expected response.
        """
        url = 'weekly_prod_report'
        response = self.client.post(reverse(url))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b'{"error": "Invalid request method"}')

    def test_request_get_weekly_prod_report(self):
        """
        Test that a GET request to the weekly_prod_report endpoint returns the expected response.
        """
        url = 'weekly_prod_report'
        response = self.client.get(reverse(url))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Disposition'], 'attachment; filename=weekly-prod-report.xlsx')

    def test_weekly_prod_report(self):
        """
        Test the weekly_prod_report function.
        """
        report = weekly_prod_report()
        self.assertEqual(report, [[('R2', 'D2', 1), ('R2', 'D3', 1)], [('R3', 'D2', 1), ('R3', 'D3', 1)]])

    def test_weekly_prod_report_to_excel(self):
        """
        Test the weekly_prod_report_to_excel function.
        """
        report = weekly_prod_report()
        workbook = weekly_prod_report_to_excel(report)
        workbook.save('test_workbook.xlsx')

        loaded_workbook = load_workbook('test_workbook.xlsx')
        worksheets = loaded_workbook.worksheets
        expected_contents = [
            ['Модель', 'Версия', 'Количество за неделю'],
            ['R2', 'D2', 1],
            ['R2', 'D3', 1],
            ['Модель', 'Версия', 'Количество за неделю'],
            ['R3', 'D2', 1],
            ['R3', 'D3', 1]
        ]
        data = []

        for worksheet in worksheets:
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))

            for source_row, test_row in zip(data, expected_contents):
                self.assertEqual(source_row, test_row)

            self.assertEqual(workbook.sheetnames, ['R2', 'R3'])
        os.remove('test_workbook.xlsx')
